"""Generate a formatted competency tracking workbook from the bundled dataset."""
from __future__ import annotations

import argparse
import json
from collections import Counter
from itertools import cycle
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_INPUT_PATH = Path("data/pgr_competency_checklist_bundle.json")
DEFAULT_OUTPUT_PATH = Path("data/PGR_Competency_Checklist.xlsx")

KNOWN_SECTION_COLOURS: Dict[str, Tuple[str, str]] = {
    "Orientation & Foundations": ("EDF2FB", "1B4F72"),
    "Outlet Competencies": ("E8F5E9", "1B5E20"),
    "Compliance & Guest Care": ("FFF4E6", "E65100"),
}

SECTION_COLOUR_FALLBACKS: Tuple[Tuple[str, str], ...] = (
    ("E8ECFF", "1F3B70"),
    ("E8F5E9", "1B5E20"),
    ("FFF4E5", "8E4B10"),
    ("FCE4EC", "6A1B4D"),
    ("F1F8E9", "33691E"),
)

STATUS_COLOUR_PRESETS = {
    "Not Started": "E0E0E0",
    "In Progress": "FFF3CD",
    "Complete": "D4EDDA",
}

STATUS_COLOUR_FALLBACKS = ("E2E8F0", "FFEFD5", "E3F2FD", "F8D7DA")

PRIMARY_TITLE_COLOUR = "0B1E3F"
SECONDARY_TEXT_COLOUR = "3D4A5C"
DIVIDER_COLOUR = "D0D7DE"
OVERVIEW_HEADER_FILL = "E9F2FF"
MILESTONE_HEADER_FILL = "0B3D91"
MILESTONE_TEXT_COLOUR = "FFFFFF"
CHECKBOX_OPTIONS = ("☐ Pending", "☑ Complete")
SIGN_OFF_OPTIONS = ("☐ Awaiting", "☑ Signed")

BORDER_THIN = Border(
    left=Side(border_style="thin", color=DIVIDER_COLOUR),
    right=Side(border_style="thin", color=DIVIDER_COLOUR),
    top=Side(border_style="thin", color=DIVIDER_COLOUR),
    bottom=Side(border_style="thin", color=DIVIDER_COLOUR),
)


def make_fill(colour: str) -> PatternFill:
    """Return a solid PatternFill for a given hex colour string."""

    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def compute_section_styles(template: dict) -> Dict[str, Tuple[str, str]]:
    """Assign fills and text colours for each template section."""

    resolved: Dict[str, Tuple[str, str]] = {}
    fallback_cycle = cycle(SECTION_COLOUR_FALLBACKS)
    for section in template.get("sections", []):
        title = section.get("title", "")
        if not title:
            continue
        if title not in resolved:
            resolved[title] = KNOWN_SECTION_COLOURS.get(title, next(fallback_cycle))
    return resolved


def compute_status_fills(status_options: Iterable[str]) -> Dict[str, PatternFill]:
    """Return PatternFill instances keyed by each possible status option."""

    fills: Dict[str, PatternFill] = {}
    fallback_cycle = cycle(STATUS_COLOUR_FALLBACKS)
    for option in status_options:
        colour = STATUS_COLOUR_PRESETS.get(option, next(fallback_cycle))
        fills[option] = make_fill(colour)
    return fills


def build_workbook(dataset: dict, template: dict) -> Workbook:
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"

    section_styles = compute_section_styles(template)
    status_options = template.get("statusOptions") or list(STATUS_COLOUR_PRESETS.keys())
    status_fills = compute_status_fills(status_options)

    build_overview_sheet(
        overview, dataset, template, section_styles, status_options, status_fills
    )
    build_matrix_sheet(
        wb, dataset, template, section_styles, status_options, status_fills
    )
    build_progress_sheet(wb, dataset, template, section_styles)
    build_area_status_board(wb, dataset, template, section_styles, status_options)

    return wb


def build_overview_sheet(
    ws,
    dataset: dict,
    template: dict,
    section_styles: Dict[str, Tuple[str, str]],
    status_options: List[str],
    status_fills: Dict[str, PatternFill],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = template["title"]
    ws["A1"].font = Font(size=22, bold=True, color=PRIMARY_TITLE_COLOUR)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = make_fill(OVERVIEW_HEADER_FILL)

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Version {template['version']} | Dataset updated {dataset['updated']}"
    ws["A2"].font = Font(size=11, color=SECONDARY_TEXT_COLOUR)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "How to use this workbook"
    ws["A4"].font = Font(size=14, bold=True, color=PRIMARY_TITLE_COLOUR)

    instructions = [
        "Review the Competency Matrix tab to track each team member's progress.",
        "Update the Status column using the drop-down menu for each competency.",
        "Record completion dates and any coaching notes for full transparency.",
        "Use the TM checklist and sign-off columns to capture verification at key milestones.",
        "Refer to the Progress Summary and Area Status Board tabs for at-a-glance completion metrics.",
    ]

    for idx, item in enumerate(instructions, start=5):
        ws[f"A{idx}"] = f"• {item}"
        ws[f"A{idx}"].alignment = Alignment(horizontal="left")
        ws[f"A{idx}"].font = Font(size=11)

    start_row = 11
    render_section_palette(ws, template, section_styles, start_row)

    legend_row = start_row + len(template.get("sections", [])) + 4
    render_status_legend(ws, status_options, status_fills, legend_row)

    timeline_row = legend_row + len(status_options) + 4
    render_milestone_timeline(ws, timeline_row)

    for column, width in {
        "A": 32,
        "B": 42,
        "C": 26,
        "D": 22,
        "E": 22,
        "F": 18,
        "G": 18,
    }.items():
        ws.column_dimensions[column].width = width


def render_section_palette(
    ws, template: dict, section_styles: Dict[str, Tuple[str, str]], start_row: int
) -> None:
    ws[f"A{start_row - 1}"] = "Area colour palette"
    ws[f"A{start_row - 1}"].font = Font(size=14, bold=True, color=PRIMARY_TITLE_COLOUR)

    headers = ("Area", "Focus", "Colour")
    for idx, header in enumerate(headers, start=0):
        col_letter = get_column_letter(idx + 1)
        cell = ws[f"{col_letter}{start_row}"]
        cell.value = header
        cell.font = Font(bold=True, color=PRIMARY_TITLE_COLOUR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = make_fill("EDF2FF")
        cell.border = BORDER_THIN

    row = start_row + 1
    for section in template.get("sections", []):
        title = section.get("title", "")
        if not title:
            continue
        fill_colour, _ = section_styles.get(title, ("F8FAFC", PRIMARY_TITLE_COLOUR))
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = Font(bold=True, color=PRIMARY_TITLE_COLOUR)
        ws[f"A{row}"].border = BORDER_THIN
        ws[f"B{row}"] = section.get(
            "description",
            "Capture notes about where this training happens and who supports it.",
        )
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{row}"].border = BORDER_THIN
        colour_cell = ws[f"C{row}"]
        colour_cell.fill = make_fill(fill_colour)
        colour_cell.border = BORDER_THIN
        row += 1


def render_status_legend(
    ws, status_options: List[str], status_fills: Dict[str, PatternFill], start_row: int
) -> None:
    ws[f"A{start_row - 1}"] = "Status legend"
    ws[f"A{start_row - 1}"].font = Font(size=14, bold=True, color=PRIMARY_TITLE_COLOUR)

    ws[f"A{start_row}"] = "Status"
    ws[f"B{start_row}"] = "Meaning"
    for cell in ws[f"A{start_row}:B{start_row}"][0]:
        cell.font = Font(bold=True, color=PRIMARY_TITLE_COLOUR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = make_fill("E9EDF5")
        cell.border = BORDER_THIN

    descriptions = {
        "Not Started": "Training not yet scheduled.",
        "In Progress": "Shadowing, buddy shifts or partial completion.",
        "Complete": "TM demonstrated competency and has been signed off.",
    }

    row = start_row + 1
    for status in status_options:
        ws[f"A{row}"] = status
        ws[f"A{row}"].border = BORDER_THIN
        ws[f"A{row}"].fill = status_fills.get(status, make_fill("FFFFFF"))
        ws[f"B{row}"] = descriptions.get(
            status,
            "Custom status used by your venue.",
        )
        ws[f"B{row}"].border = BORDER_THIN
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1


def render_milestone_timeline(ws, start_row: int) -> None:
    ws[f"A{start_row - 1}"] = "Milestone timeline"
    ws[f"A{start_row - 1}"].font = Font(size=14, bold=True, color=PRIMARY_TITLE_COLOUR)

    headers = ("Milestone", "Focus", "Owner", "Evidence")
    for idx, header in enumerate(headers, start=0):
        col_letter = get_column_letter(idx + 1)
        cell = ws[f"{col_letter}{start_row}"]
        cell.value = header
        cell.font = Font(bold=True, color=MILESTONE_TEXT_COLOUR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = make_fill(MILESTONE_HEADER_FILL)
        cell.border = BORDER_THIN

    milestones = [
        ("First shift", "Buddy shift, venue tour and basics", "Trainer", "TM checklist updated"),
        ("Week 2", "Core tasks performed with supervision", "Mentor", "Coaching notes captured"),
        ("30 days", "Confident in primary outlet", "Venue Manager", "TM sign-off"),
        ("60 days", "Cross-trained in secondary areas", "Venue Manager", "Manager sign-off"),
        ("90 days", "Ready for independent shifts", "Leadership", "Celebration + follow-up"),
    ]

    row = start_row + 1
    for milestone, focus, owner, evidence in milestones:
        ws[f"A{row}"] = milestone
        ws[f"B{row}"] = focus
        ws[f"C{row}"] = owner
        ws[f"D{row}"] = evidence
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
def build_matrix_sheet(
    wb: Workbook,
    dataset: dict,
    template: dict,
    section_styles: Dict[str, Tuple[str, str]],
    status_options: Iterable[str],
    status_fills: Dict[str, PatternFill],
) -> None:
    status_options = list(status_options)
    ws = wb.create_sheet(title="Competency Matrix")
    ws.sheet_view.showGridLines = False
    headers = [
        "Team Member",
        "Staff ID",
        "Role",
        "Assigned Mentor",
        "Start Date",
        "Area",
        "Competency",
        "Details",
        "Status",
        "TM Checklist",
        "Completed On",
        "Coaching Notes",
        "Trainer",
        "TM Sign-Off",
        "Manager Sign-Off",
        "Follow-up Date",
    ]

    ws.append(headers)
    for cell in ws["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN

    template_defaults = dataset.get("defaults", {})
    default_status = template_defaults.get("defaultStatus") or (
        status_options[0] if status_options else "Not Started"
    )

    row_idx = 2
    for person in dataset.get("people", []):
        meta_notes = person.get("notes", "")
        for section in template["sections"]:
            section_title = section["title"]
            fill_colour, text_colour = section_styles.get(
                section_title, ("F8FAFC", "1F2933")
            )
            section_fill = make_fill(fill_colour)
            for item in section["items"]:
                competency_data = person.get("competencies", {}).get(item["id"], {})
                status = competency_data.get("status")
                if not status:
                    status = item.get("defaultStatus") or default_status
                completed_on = competency_data.get("completedOn", "")
                item_notes = competency_data.get("notes", "")

                ws.append(
                    [
                        person.get("name", ""),
                        person.get("staffId", ""),
                        person.get("role", template_defaults.get("role", "")),
                        person.get("mentor", template_defaults.get("mentor", "")),
                        person.get("startDateDisplay", template_defaults.get("startDateDisplay", "")),
                        section_title,
                        item["label"],
                        item.get("description", ""),
                        status,
                        CHECKBOX_OPTIONS[0],
                        completed_on,
                        build_notes(meta_notes, item_notes),
                        competency_data.get("trainer", ""),
                        "",
                        "",
                        "",
                    ]
                )

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = BORDER_THIN
                    if col_idx == 6:
                        cell.fill = section_fill
                        cell.font = Font(bold=True, color=text_colour)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx == 8:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    elif col_idx == 9:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx == 12:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    elif col_idx in (10, 13, 14, 15):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_idx in (11, 16):
                        cell.number_format = "dd mmm yyyy"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center")
                row_idx += 1

    column_widths = {
        "A": 22,
        "B": 12,
        "C": 20,
        "D": 20,
        "E": 14,
        "F": 24,
        "G": 34,
        "H": 48,
        "I": 16,
        "J": 14,
        "K": 16,
        "L": 32,
        "M": 18,
        "N": 18,
        "O": 18,
        "P": 16,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    data_range = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"
    ws.auto_filter.ref = data_range

    for r in range(2, row_idx):
        ws.row_dimensions[r].height = 24

    if row_idx == 2:
        return

    status_range = f"I2:I{row_idx - 1}"
    status_validation = DataValidation(
        type="list",
        formula1="\"" + ",".join(status_options) + "\"",
        allow_blank=True,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(status_range)

    checklist_range = f"J2:J{row_idx - 1}"
    checklist_validation = DataValidation(
        type="list",
        formula1="\"" + ",".join(CHECKBOX_OPTIONS) + "\"",
        allow_blank=True,
    )
    ws.add_data_validation(checklist_validation)
    checklist_validation.add(checklist_range)

    tm_signoff_range = f"N2:N{row_idx - 1}"
    manager_signoff_range = f"O2:O{row_idx - 1}"
    signoff_validation = DataValidation(
        type="list",
        formula1="\"" + ",".join(SIGN_OFF_OPTIONS) + "\"",
        allow_blank=True,
    )
    ws.add_data_validation(signoff_validation)
    signoff_validation.add(tm_signoff_range)
    signoff_validation.add(manager_signoff_range)

    for date_range in (f"K2:K{row_idx - 1}", f"P2:P{row_idx - 1}"):
        date_validation = DataValidation(
            type="date",
            operator="greaterThan",
            formula1="DATE(2020,1,1)",
            allow_blank=True,
        )
        ws.add_data_validation(date_validation)
        date_validation.add(date_range)

    for status, fill in status_fills.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=["\"" + status + "\""], fill=fill),
        )


def build_progress_sheet(
    wb: Workbook, dataset: dict, template: dict, section_styles: Dict[str, Tuple[str, str]]
) -> None:
    ws = wb.create_sheet(title="Progress Summary")
    ws.sheet_view.showGridLines = False

    headers = ["Team Member"]
    headers.extend(section["title"] for section in template["sections"])
    headers.extend(["Not Started", "In Progress", "Complete", "Overall %", "Last Updated"])
    ws.append(headers)
    for col_idx, cell in enumerate(ws["1:1"], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        if 1 < col_idx <= len(template["sections"]) + 1:
            section = template["sections"][col_idx - 2]
            fill_colour, text_colour = section_styles.get(
                section["title"], ("1F2933", "FFFFFF")
            )
            cell.fill = make_fill(fill_colour)
            cell.font = Font(bold=True, color=text_colour)
        else:
            cell.fill = PatternFill(
                start_color="334155", end_color="334155", fill_type="solid"
            )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN

    updated = dataset.get("updated", "")
    for person in dataset.get("people", []):
        row = [person.get("name", "")]
        total_completed = 0
        total_items = 0
        status_totals = Counter()
        for section in template["sections"]:
            items = section["items"]
            completed = 0
            for item in items:
                competency = person.get("competencies", {}).get(item["id"], {})
                status = competency.get("status")
                if status:
                    status_totals[status] += 1
                if status == "Complete":
                    completed += 1
            total_items += len(items)
            total_completed += completed
            row.append(f"{completed} / {len(items)}")

        overall_percent = round((total_completed / total_items) * 100) if total_items else 0
        row.extend([
            status_totals.get("Not Started", 0),
            status_totals.get("In Progress", 0),
            status_totals.get("Complete", 0),
        ])
        row.append(f"{overall_percent}%")
        row.append(updated)
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    if ws.max_row >= 2:
        percent_col_letter = get_column_letter(len(headers) - 1)
        percent_range = f"{percent_col_letter}2:{percent_col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            percent_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["90"],
                fill=PatternFill(
                    start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
                ),
            ),
        )
        ws.conditional_formatting.add(
            percent_range,
            CellIsRule(
                operator="between",
                formula=["60", "89"],
                fill=PatternFill(
                    start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                ),
            ),
        )
        ws.conditional_formatting.add(
            percent_range,
            CellIsRule(
                operator="lessThan",
                formula=["60"],
                fill=PatternFill(
                    start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
                ),
            ),
        )

    column_widths = {
        "A": 26,
        "B": 22,
        "C": 22,
        "D": 22,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def build_area_status_board(
    wb: Workbook,
    dataset: dict,
    template: dict,
    section_styles: Dict[str, Tuple[str, str]],
    status_options: List[str],
) -> None:
    ws = wb.create_sheet(title="Area Status Board")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Area status board"
    ws["A1"].font = Font(size=18, bold=True, color=PRIMARY_TITLE_COLOUR)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = make_fill("EFF6FF")

    headers = [
        "Area",
        "Total Items",
        "Not Started",
        "In Progress",
        "Complete",
        "Other",
        "Completion %",
        "Colour",
    ]
    ws.append(headers)
    for cell in ws["2:2"]:
        cell.font = Font(bold=True, color=PRIMARY_TITLE_COLOUR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = make_fill("E5ECF8")
        cell.border = BORDER_THIN

    template_defaults = dataset.get("defaults", {})
    default_status = template_defaults.get("defaultStatus") or (
        status_options[0] if status_options else "Not Started"
    )

    row = 3
    for section in template.get("sections", []):
        title = section.get("title", "")
        fill_colour, text_colour = section_styles.get(
            title, ("F8FAFC", PRIMARY_TITLE_COLOUR)
        )
        counts = Counter()
        for person in dataset.get("people", []):
            competencies = person.get("competencies", {})
            for item in section.get("items", []):
                competency = competencies.get(item.get("id", ""), {})
                status = competency.get("status")
                if not status:
                    status = item.get("defaultStatus") or default_status or "Not Started"
                counts[status] += 1

        total = sum(counts.values())
        complete = counts.get("Complete", 0)
        not_started = counts.get("Not Started", 0)
        in_progress = counts.get("In Progress", 0)
        other = total - (complete + not_started + in_progress)
        completion_percent = round((complete / total) * 100) if total else 0

        ws.append(
            [
                title,
                total,
                not_started,
                in_progress,
                complete,
                other,
                f"{completion_percent}%",
                "",
            ]
        )

        ws[f"A{row}"].font = Font(bold=True, color=text_colour)
        ws[f"A{row}"].fill = make_fill(fill_colour)
        ws[f"H{row}"].fill = make_fill(fill_colour)
        ws[f"H{row}"] = ""

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = BORDER_THIN
            if col_idx == 7:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    ws.freeze_panes = "A3"

    if row == 3:
        return

    ws.auto_filter.ref = f"A2:H{row - 1}"

    for r in range(3, row):
        ws.row_dimensions[r].height = 22

    column_widths = {
        "A": 30,
        "B": 14,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 14,
        "G": 16,
        "H": 12,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    data_ref = Reference(ws, min_col=3, max_col=5, min_row=2, max_row=row - 1)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=row - 1)
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "Progress by area"
    chart.y_axis.title = "Competencies"
    chart.x_axis.title = "Area"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, f"A{row + 1}")
def build_notes(meta_notes: str, item_notes: str) -> str:
    notes = []
    if meta_notes:
        notes.append(meta_notes)
    if item_notes:
        notes.append(item_notes)
    return "\n\n".join(notes)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a formatted competency tracking workbook."
    )
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help="Path to the competency bundle (JSON or TXT).",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Destination for the generated Excel workbook.",
    )
    return parser.parse_args()


def load_bundle(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"Unable to locate competency bundle at {path}")

    with path.open("r", encoding="utf-8") as handle:
        try:
            payload = json.load(handle)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{path} is not valid JSON: {exc}") from exc

    missing_keys = {"dataset", "template"} - payload.keys()
    if missing_keys:
        raise KeyError(
            f"Bundle at {path} is missing required keys: {', '.join(sorted(missing_keys))}"
        )

    return payload


def main() -> None:
    args = parse_args()
    bundle_path: Path = args.input

    if not bundle_path.exists() and bundle_path == DEFAULT_INPUT_PATH:
        legacy_path = Path("data/(5) PGR Competency Checklist.txt")
        if legacy_path.exists():
            bundle_path = legacy_path

    payload = load_bundle(bundle_path)

    dataset = payload["dataset"]
    template = payload["template"]

    workbook = build_workbook(dataset, template)

    output_path: Path = args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Workbook written to {output_path}")


if __name__ == "__main__":
    main()
